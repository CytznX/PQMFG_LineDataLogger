import openpyxl
import os


def write(text, path = 'hello' ):
	fullPath = path+'.xls'


	#Usefull commands
	#wb.get_sheet_names() #<--- gets all sheet names in WB

	if not os.path.exists(fullPath):
		wb = openpyxl.Workbook()
		ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
		ws.title = "LogSummary"
		ws.cell('A1').value = text
		ws.cell('A2').value = 'hahahah'
		wb.save(fullPath)

	else:

		wb = openpyxl.load_workbook(fullPath)
		
		#wb.get_sheet_by_name("LogSummary")
		ws = wb.get_active_sheet()

		ws.cell(row = len(ws.rows), column = len(ws.rows[-1])+1).value=text
		wb.save(fullPath)



	

write(str(raw_input('>> ')))