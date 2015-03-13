#!/usr/bin/env python

import sys

import socket, os
import datetime, time
import pickle

import re
import MySQLdb
import _mysql

from threading import Thread
from openpyxl import *
from openpyxl import styles
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import Border, Side

class ThreadedTCPNetworkAgent(Thread):

	'''Default constructor... Much Love'''
	def __init__(self, portNum, Folder='WorkOrderExcelLogs', BuffSize=1024):

		#Initialize myself as thread... =P
		Thread.__init__(self)

		#setup some class variables
		self.running = True
		self.DefaultClientPort = 5005
		self._BuffSize = BuffSize
		self.Addr = ('', portNum)

		self.CurrentLines = dict()

		#self.IsMachineAlive = Thread(target=self.whoIsAlive, args=())
		#self.IsMachineAlive.start()

		#FOR TESTING ONLY
		self.testing = True
		'''if self.testing:
			self.CurrentLines[3] =("192.168.20.198", 5005)
			self.CurrentLines[4] =("192.168.20.198", 5005)
			self.CurrentLines[5] =("192.168.20.198", 5005)
			self.CurrentLines[11] =("192.168.20.198", 5005)
			self.CurrentLines[34] =("192.168.20.198", 5005)
			self.CurrentLines[7] =("192.168.20.198", 5005)
		'''

		#Creates the logg folder if it doesnt exist
		self.WO_LogFolder = Folder+'/'
		if not os.path.isdir(self.WO_LogFolder):
			os.makedirs(self.WO_LogFolder)

		#create the socket that will be listening on
		self.serversock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
		self.serversock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
		self.serversock.bind(self.Addr)
		self.serversock.listen(5)

	def whoIsAlive(self):
		while self.running:
			for key in self.CurrentLines.keys():

				try:
					Test = socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect(self.CurrentLines[key])
					Test.send("#ALIVE")
					inData = Test.recv(self._BuffSize)
					if not inData:
						del(self.CurrentLines[key])

				except socket.error:
					del(self.CurrentLines[key])

				except TypeError:
					print "UHHHHHHH Typer error?", key, self.DefaultClientPort

			time.sleep(60)#Delay for 1 minute....


	'''Heres where we spawn a minin thread that manages a individual connection to this machine'''
	def miniThread(self,clientsock,addr):

		#Notify Connection made
		#print 'Connection Recieved@'+datetime.datetime.now().strftime('(%D, %H:%M:%S)')+' From Addr: '+str(addr)

		#the collected message
		safety = ''
		data = ''
		can_continue = False
		now = datetime.datetime.now()

		#PULL IN NEW
		while True:
			data = clientsock.recv(self._BuffSize)

			#Heres we check to see if we have any data
			if not data:
				if safety is not "":
					can_continue = True
				break

			#This is the kill switch
			elif data.startswith("#KILL") or (safety+data).startswith("#KILL"):
				self.stop()
				break

			safety += data

		if can_continue:
			print "Processing Connection", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

			try:
				unPickledData = pickle.loads(safety)
				#print type(unPickledData), len(unPickledData)
				self.writeToSQL(MachineLog=unPickledData)
				self.writeToExel(unPickledData, self.WO_LogFolder)

			except IndexError, e:
				print "index error??????\n", e
		# except KeyError,e:
		# 	print "fucking key error\n", e

	def writeToExel(self, unPickledData, WO_LogFolder):

		now = datetime.datetime.now()

		if not os.path.exists(WO_LogFolder):
			os.makedirs(WO_LogFolder)

		w0 = unPickledData[0]["WO"]
		if not os.path.isfile(WO_LogFolder + str(w0) + ".xlsx"):

			# Create a new
			wb = Workbook()
			headerSheet = wb.active
			curNumRuns = 0

			headerSheet.title = "Work Order Sumary"
			headerSheet['A1'] = "WO#"
			headerSheet['B1'] = "Product Name"

			headerSheet['A1'].style = styles.Style(font=Font(size=15, bold=True), border=Border(bottom=Side(style='thick'), right=Side(style='thin')), alignment=Alignment(horizontal="center"))
			headerSheet['B1'].style = styles.Style(font=Font(size=15, bold=True), border=Border(bottom=Side(style='thick')), alignment=Alignment(horizontal="center"))

			headerSheet['A2'] = w0
			try:
				headerSheet['B2'] = unPickledData[0]["Item Number"]
			except KeyError, e:
				headerSheet['B2'] = "N/A"

			headerSheet['A2'].style = styles.Style(border=Border(right=Side(style='thin')), alignment=Alignment(horizontal="center"))
			headerSheet['B2'].style = styles.Style(alignment=Alignment(horizontal="center"))

			headerSheet.merge_cells('A4:B4')
			headerSheet["A4"] = "Work Order Totals"

			headerSheet["A4"].style = styles.Style(font=Font(size=15, bold=True), border=Border(bottom=Side(style='thick')), alignment=Alignment(horizontal="center"))

			workOrderTotalHeaders = ["WO Start", "WO End", "Runtime", "Count", "Scrap", "Run Rate(Pcs/Hr)", "Cost/Peace"]

			for row, header in enumerate(workOrderTotalHeaders):
				headerSheet['A' + str(5 + row)] = header + ":"
				headerSheet['A' + str(5 + row)].style = styles.Style(border=Border(right=Side(style='thin')), alignment=Alignment(horizontal="center"))

			headerSheet['B5'] = unPickledData[0]["WO StartTime"].strftime('%H:%M:%S (%D)')
			headerSheet['B5'].style = styles.Style(alignment=Alignment(horizontal="center"))

			headerSheet['B6'] = unPickledData[0]["Time Log Created"].strftime('%H:%M:%S (%D)')
			headerSheet['B6'].style = styles.Style(alignment=Alignment(horizontal="center"))

			headerSheet['B7'] = str(unPickledData[0]["Time Log Created"] - unPickledData[0]["WO StartTime"]).split(".")[0]
			headerSheet['B7'].style = styles.Style(alignment=Alignment(horizontal="center"))

			headerSheet['B8'] = sum(unPickledData[0]["Total Count"])
			headerSheet['B8'].style = styles.Style(alignment=Alignment(horizontal="center"))

			headerSheet['B9'] = unPickledData[0]["Fail Count"]
			headerSheet['B9'].style = styles.Style(alignment=Alignment(horizontal="center"))

			try:
				headerSheet['B10'] = sum(unPickledData[0]["Total Count"]) / (unPickledData[0]["Fill End"] - unPickledData[0]["Fill Start"]).seconds / 60.0 / 60.0
			except TypeError:
				headerSheet['B10'] = "N/A"
			finally:
				headerSheet['B10'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['B11'] = (sum(unPickledData[0]["Total Count"]) - unPickledData[0]["Fail Count"]) / (unPickledData[0]["Time Log Created"] - unPickledData[0]["WO StartTime"]).seconds
			headerSheet['B11'].style = styles.Style(alignment=Alignment(horizontal="center"))

			headerSheet['B11'] = "No Formula"
			headerSheet['B11'].style = styles.Style(alignment=Alignment(horizontal="center"))

			headerSheet.merge_cells('A' + str(6 + len(workOrderTotalHeaders)) + ':B' + str(6 + len(workOrderTotalHeaders)))
			headerSheet["A" + str(6 + len(workOrderTotalHeaders))] = "Pallet Totals"
			headerSheet["A" + str(6 + len(workOrderTotalHeaders))].style = styles.Style(font=Font(size=15, bold=True), border=Border(bottom=Side(style='thick')), alignment=Alignment(horizontal="center"))

			headerSheet["A" + str(7 + len(workOrderTotalHeaders))] = "# of Pallets:"
			headerSheet["A" + str(7 + len(workOrderTotalHeaders))].style = styles.Style(border=Border(right=Side(style='thin')), alignment=Alignment(horizontal="center"))

			headerSheet["A" + str(8 + len(workOrderTotalHeaders))] = "Peaces Count:"
			headerSheet["A" + str(8 + len(workOrderTotalHeaders))].style = styles.Style(border=Border(right=Side(style='thin')), alignment=Alignment(horizontal="center"))

			headerSheet.merge_cells('A' + str(10 + len(workOrderTotalHeaders)) + ':B' + str(10 + len(workOrderTotalHeaders)))
			headerSheet["A" + str(10 + len(workOrderTotalHeaders))] = "Downtime Totals"
			headerSheet["A" + str(10 + len(workOrderTotalHeaders))].style = styles.Style(font=Font(size=15, bold=True), border=Border(bottom=Side(style='thick')), alignment=Alignment(horizontal="center"))

			workOrderDownTimeHeaders = ["Maintenance", "Inventory", "QC", "Break", "Change Over"," Total"]

			for row, header in enumerate(workOrderDownTimeHeaders):
				headerSheet['A' + str(11 + len(workOrderTotalHeaders) + row)] = header + ":"
				headerSheet['A' + str(11 + len(workOrderTotalHeaders) + row)].style = styles.Style(border=Border(right=Side(style='thin')), alignment=Alignment(horizontal="center"))

			peacesSums = 0
			palletSums = len(sorted(unPickledData[5], key=unPickledData[4].get)[:-1])
			for palletInfoItems in sorted(unPickledData[5], key=unPickledData[4].get)[:-1]:
				try:
					peacesSums += int(unPickledData[5][palletInfoItems][3])
				except ValueError:
					pass


			headerSheet['B14'] = palletSums
			headerSheet['B14'].style = styles.Style(alignment=Alignment(horizontal="center"))

			headerSheet['B15'] = peacesSums
			headerSheet['B15'].style = styles.Style(alignment=Alignment(horizontal="center"))

			for count, formatedTime in enumerate(["FormattedMain", "FormattedInv", "FormattedQuality", "FormattedBreak", "FormattedChngOvr", "FormattedTotal"]):
				headerSheet["B" + str(18 + count)] = "%0.2d:%0.2d:%0.2d" % (unPickledData[2][formatedTime][0], unPickledData[2][formatedTime][1], unPickledData[2][formatedTime][2])
				headerSheet["B" + str(18 + count)].style = styles.Style(alignment=Alignment(horizontal="center"))

			# Creates Headers For Data Columbs

			RunInfo_Headers = ["Run#", "Line#", "Job Start", "Fill Start", "Fill End", "Job End", "Runtime", "Count", "Pallets", "Scrap", "Downtime", "Run Rate (Pcs/Hr)", "# of Line Workers", "Cost/Pcs"]
			CharStart = "D"
			HeaderLine = 4

			for header in RunInfo_Headers:
				headerSheet[CharStart + str(HeaderLine)] = header
				headerSheet[CharStart + str(HeaderLine)].style = styles.Style(font=Font(bold=True), border=Border(bottom=Side(style='thick'), right=Side(style='thin')), alignment=Alignment(horizontal="center"))
				CharStart = chr(ord(CharStart) + 1)

			headerSheet['D5'] = "1"
			headerSheet['D5'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['E5'] = unPickledData[0]["Machine ID"]
			headerSheet['E5'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['F5'] = unPickledData[0]["WO StartTime"].strftime('%H:%M:%S (%D)')
			headerSheet['F5'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			try:
				headerSheet['G5'] = unPickledData[0]["Fill Start"].strftime('%H:%M:%S (%D)')
			except AttributeError:
				headerSheet['G5'] = "N/A"
			finally:
				headerSheet['G5'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			try:
				headerSheet['H5'] = unPickledData[0]["Fill End"].strftime('%H:%M:%S (%D)')
			except AttributeError:
				headerSheet['H5'] = "N/A"
			finally:
				headerSheet['H5'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['I5'] = unPickledData[0]["Time Log Created"].strftime('%H:%M:%S (%D)')
			headerSheet['I5'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			try:
				headerSheet['J5'] = str(unPickledData[0]["Fill End"] - unPickledData[0]["Fill Start"]).split(".")[0]
			except:
				headerSheet['J5'] = "N/A"
			finally:
				headerSheet['J5'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['K5'] = sum(unPickledData[0]["Total Count"])
			headerSheet['K5'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['L5'] = len(unPickledData[5]) - 1
			headerSheet['L5'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['M5'] = unPickledData[0]["Fail Count"]
			headerSheet['M5'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['N5'] = "%0.2d:%0.2d:%0.2d" % (unPickledData[2]["FormattedTotal"][0], unPickledData[2]["FormattedTotal"][1], unPickledData[2]["FormattedTotal"][2])
			headerSheet['N5'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			try:
				headerSheet['O5'] = sum(unPickledData[0]["Total Count"]) / (unPickledData[0]["Fill End"] - unPickledData[0]["Fill Start"]).seconds / 60.0 / 60.0
			except TypeError:
				headerSheet['O5'] = "N/A"
			finally:
				headerSheet['O5'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['P5'] = len(unPickledData[1].keys())
			headerSheet['P5'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['Q5'] = "No Formula"
			headerSheet['Q5'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet.column_dimensions["A"].width = 30.0
			headerSheet.column_dimensions["B"].width = 30.0

			headerSheet.column_dimensions["D"].width = 15.0
			headerSheet.column_dimensions["E"].width = 30.0
			headerSheet.column_dimensions["F"].width = 30.0
			headerSheet.column_dimensions["G"].width = 15.0
			headerSheet.column_dimensions["H"].width = 15.0
			headerSheet.column_dimensions["I"].width = 15.0

			headerSheet.row_dimensions[1].height = 20
			headerSheet.row_dimensions[4].height = 20

		else:

			wb = load_workbook(WO_LogFolder + str(w0) + '.xlsx')

			headerSheet = wb.get_sheet_by_name("Work Order Sumary")

			curNumRuns = len(wb.get_sheet_names()) - 1

			headerSheet['B6'].value = unPickledData[0]["Time Log Created"].strftime('%H:%M:%S (%D)')
			headerSheet['B6'].style = styles.Style(alignment=Alignment(horizontal="center"))

			FormerVal = headerSheet['B7'].value.split(":")
			curVal = str(unPickledData[0]["Time Log Created"] - unPickledData[0]["WO StartTime"]).split(".")[0].split(":")

			sec = int(FormerVal[2]) + int(curVal[2])
			minutes = int(FormerVal[1]) + int(curVal[1])
			hours = int(FormerVal[0]) + int(curVal[0])

			if sec >= 60:
				minutes += 1
				sec = sec - 60

			if minutes >= 60:
				hours += 1
				minutes = minutes - 60

			headerSheet['B7'] = str(hours) + ":" + str(minutes) + ":" + str(sec)
			headerSheet['B7'].style = styles.Style(alignment=Alignment(horizontal="center"))

			formerVal = headerSheet['B8'].value

			headerSheet['B8'] = int(formerVal) + sum(unPickledData[0]["Total Count"])
			headerSheet['B8'].style = styles.Style(alignment=Alignment(horizontal="center"))

			formerVal = headerSheet['B9'].value

			headerSheet['B9'] = int(formerVal) + unPickledData[0]["Fail Count"]
			headerSheet['B9'].style = styles.Style(alignment=Alignment(horizontal="center"))

			if headerSheet['B10'].value == "N/A":
				try:
					headerSheet['B10'] = sum(unPickledData[0]["Total Count"]) / (unPickledData[0]["Fill End"] - unPickledData[0]["Fill Start"]).seconds / 60.0 / 60.0
				except TypeError:
					headerSheet['B10'] = "N/A"
				finally:
					headerSheet['B10'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			else:

				try:
					headerSheet['B10'] = int(headerSheet['B10'].value) + sum(unPickledData[0]["Total Count"]) / (unPickledData[0]["Fill End"] - unPickledData[0]["Fill Start"]).seconds / 60.0 / 60.0
				except TypeError:
					headerSheet['B10'] = "N/A"
				finally:
					headerSheet['B10'].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			peacesSums = 0
			palletSums = len(sorted(unPickledData[5], key=unPickledData[4].get)[:-1])
			for palletInfoItems in sorted(unPickledData[5], key=unPickledData[4].get)[:-1]:
				peacesSums += int(unPickledData[5][palletInfoItems][3])

			headerSheet['B14'] = int(headerSheet['B14'].value) + palletSums
			headerSheet['B14'].style = styles.Style(alignment=Alignment(horizontal="center"))

			headerSheet['B15'] = int(headerSheet['B15'].value) + peacesSums
			headerSheet['B15'].style = styles.Style(alignment=Alignment(horizontal="center"))

			for count, formatedTime in enumerate(["FormattedMain", "FormattedInv", "FormattedQuality", "FormattedBreak", "FormattedChngOvr", "FormattedTotal"]):

				FormerVal = headerSheet["B" + str(18 + count)].value.split(":")
				curVal = (unPickledData[2][formatedTime][0], unPickledData[2][formatedTime][1], unPickledData[2][formatedTime][2])
				sec = int(FormerVal[2]) + int(curVal[2])
				minutes = int(FormerVal[1]) + int(curVal[1])
				hours = int(FormerVal[0]) + int(curVal[0])

				if sec >= 60:
					minutes += 1
					sec = sec - 60

				if minutes >= 60:
					hours += 1
					minutes = minutes - 60

				headerSheet["B" + str(18 + count)] = "%0.2d:%0.2d:%0.2d" % (hours, minutes, sec)
				headerSheet["B" + str(18 + count)].style = styles.Style(alignment=Alignment(horizontal="center"))

			headerSheet['D' + str(5 + curNumRuns)] = str(curNumRuns + 1)
			headerSheet['D' + str(5 + curNumRuns)].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['E' + str(5 + curNumRuns)] = unPickledData[0]["Machine ID"]
			headerSheet['E' + str(5 + curNumRuns)].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['F' + str(5 + curNumRuns)] = unPickledData[0]["WO StartTime"].strftime('%H:%M:%S (%D)')
			headerSheet['F' + str(5 + curNumRuns)].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			try:
				headerSheet['G' + str(5 + curNumRuns)] = unPickledData[0]["Fill Start"].strftime('%H:%M:%S (%D)')
			except AttributeError:
				headerSheet['G' + str(5 + curNumRuns)] = "N/A"
			finally:
				headerSheet['G' + str(5 + curNumRuns)].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			try:
				headerSheet['H' + str(5 + curNumRuns)] = unPickledData[0]["Fill End"].strftime('%H:%M:%S (%D)')
			except AttributeError:
				headerSheet['H' + str(5 + curNumRuns)] = "N/A"
			finally:
				headerSheet['H' + str(5 + curNumRuns)].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['I' + str(5 + curNumRuns)] = unPickledData[0]["Time Log Created"].strftime('%H:%M:%S (%D)')
			headerSheet['I' + str(5 + curNumRuns)].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			try:
				headerSheet['J' + str(5 + curNumRuns)] = str(unPickledData[0]["Fill End"] - unPickledData[0]["Fill Start"]).split(".")[0]
			except:
				headerSheet['J' + str(5 + curNumRuns)] = "N/A"
			finally:
				headerSheet['J' + str(5 + curNumRuns)].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['K' + str(5 + curNumRuns)] = sum(unPickledData[0]["Total Count"])
			headerSheet['K' + str(5 + curNumRuns)].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['L' + str(5 + curNumRuns)] = len(unPickledData[5]) - 1
			headerSheet['L' + str(5 + curNumRuns)].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['M' + str(5 + curNumRuns)] = unPickledData[0]["Fail Count"]
			headerSheet['M' + str(5 + curNumRuns)].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['N' + str(5 + curNumRuns)] = "%0.2d:%0.2d:%0.2d" % (unPickledData[2]["FormattedTotal"][0], unPickledData[2]["FormattedTotal"][1], unPickledData[2]["FormattedTotal"][2])
			headerSheet['N' + str(5 + curNumRuns)].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			try:
				headerSheet['O' + str(5 + curNumRuns)] = sum(unPickledData[0]["Total Count"]) / (unPickledData[0]["Fill End"] - unPickledData[0]["Fill Start"]).seconds / 60.0 / 60.0
			except TypeError:
				headerSheet['O' + str(5 + curNumRuns)] = "N/A"
			finally:
				headerSheet['O' + str(5 + curNumRuns)].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['P' + str(5 + curNumRuns)] = len(unPickledData[1].keys())
			headerSheet['P' + str(5 + curNumRuns)].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

			headerSheet['Q' + str(5 + curNumRuns)] = "No Formula"
			headerSheet['Q' + str(5 + curNumRuns)].style = styles.Style(alignment=Alignment(horizontal="center"), border=Border(right=Side(style='thin')))

		# -------------------------------------------------------------------------------
		# --------------------------------RUN-SHEET-HEADER-------------------------------
		# -------------------------------------------------------------------------------

		_LastColumb = 1

		_preOrderedKeys = ["Machine ID", "WO", "Bulk Wo", "WO StartTime", "Time Log Created", "Total Count", "Box Count", "Fail Count", "Peaces Per Box", "Fill Start", "Fill End"]
		_TimedKeys = ["WO StartTime", "Time Log Created", "Fill Start", "Fill End"]

		FirstSheet = wb.create_sheet()
		FirstSheet.title = 'Run#' + str(curNumRuns + 1)

		FirstSheet['A' + str(_LastColumb)] = "----Run #" + str(curNumRuns + 1) + " Data----"
		FirstSheet['A' + str(_LastColumb)].style = styles.Style(font=Font(bold=True), alignment=Alignment(horizontal="left"))

		_LastColumb += 1

		for key in _preOrderedKeys:
			if key not in _TimedKeys:

				FirstSheet['A' + str(_LastColumb)] = key + ":"
				FirstSheet['A' + str(_LastColumb)].style = styles.Style(font=Font(bold=True), alignment=Alignment(horizontal="center"))

				try:
					FirstSheet['B' + str(_LastColumb)] = unPickledData[0][key]
					FirstSheet['B' + str(_LastColumb)].style = styles.Style(alignment=Alignment(horizontal="center"))

				except KeyError, e:
					print "Woopc... Passed dictionary didnt contain Key: ", key

				except ValueError, e:

					FirstSheet['B' + str(_LastColumb)] = sum(unPickledData[0][key])
					FirstSheet['B' + str(_LastColumb)].style = styles.Style(alignment=Alignment(horizontal="center"))
					_LastColumb += 1

					FirstSheet['A' + str(_LastColumb)] = "^^^Hrly"
					FirstSheet['A' + str(_LastColumb)].style = styles.Style(alignment=Alignment(horizontal="center"))

					FirstSheet['B' + str(_LastColumb)] = str(unPickledData[0][key])
					FirstSheet['B' + str(_LastColumb)].style = styles.Style(alignment=Alignment(horizontal="center"))

			else:
				try:
					FirstSheet['A' + str(_LastColumb)] = key + ":"
					FirstSheet['A' + str(_LastColumb)].style = styles.Style(font=Font(bold=True), alignment=Alignment(horizontal="center"))

					if not unPickledData[0][key] == None:
						FirstSheet['B' + str(_LastColumb)] = unPickledData[0][key].strftime('%H:%M:%S (%D)')
					else:
						FirstSheet['B' + str(_LastColumb)] = "N/A"
					FirstSheet['B' + str(_LastColumb)].style = styles.Style(alignment=Alignment(horizontal="center"))
				except:
					print "uh oh something went wrong...", key

			_LastColumb += 1
		_LastColumb += 1

		FirstSheet['A' + str(_LastColumb)].style = styles.Style(border=Border(top=Side(style='thick')))
		FirstSheet['B' + str(_LastColumb)].style = styles.Style(border=Border(top=Side(style='thick')))
		FirstSheet['C' + str(_LastColumb)].style = styles.Style(border=Border(top=Side(style='thick')))

		_LastColumb += 1

		#-------------------------------------------------------------------------------
		#--------------------------------FILLSHEET--------------------------------------
		#-------------------------------------------------------------------------------

		FirstSheet['A'+str(_LastColumb)] = "----Fillsheet Data----"
		FirstSheet['A'+str(_LastColumb)].style = styles.Style(font=Font(bold=True, ), alignment=Alignment(horizontal="left"))
		_LastColumb += 1

		for fillSheetItems in unPickledData[3].keys():

			#Set Data
			FirstSheet['A'+str(_LastColumb)] = fillSheetItems+":"
			FirstSheet['B'+str(_LastColumb)] = unPickledData[3][fillSheetItems]

			#Format
			FirstSheet['A'+str(_LastColumb)].style = styles.Style(font=Font(bold=True, ), alignment=Alignment(horizontal="center"))
			FirstSheet['B'+str(_LastColumb)].style = styles.Style(alignment=Alignment(horizontal="center"))

			_LastColumb += 1
		_LastColumb += 1

		FirstSheet['A' + str(_LastColumb)].style = styles.Style(border=Border(top=Side(style='thick')))
		FirstSheet['B' + str(_LastColumb)].style = styles.Style(border=Border(top=Side(style='thick')))
		FirstSheet['C' + str(_LastColumb)].style = styles.Style(border=Border(top=Side(style='thick')))

		_LastColumb += 1


		#-------------------------------------------------------------------------------
		#--------------------------------ADJUSTMENTS------------------------------------
		#-------------------------------------------------------------------------------

		FirstSheet['A'+str(_LastColumb)] = '----Adjustments----'
		FirstSheet['A'+str(_LastColumb)].style = styles.Style(font=Font(bold=True))
		_LastColumb += 1

		if not unPickledData[0]["Line Var Adjustments"] == []:
			for adjCounts in unPickledData[0]["Line Var Adjustments"]:
				FirstSheet['A'+str(_LastColumb)] = '('+str(adjCounts[0])+', '+str(adjCounts[1])+', '+str(adjCounts[2])+', '+adjCounts[3].strftime('%H:%M:%S')+')'
				_LastColumb += 1
			_LastColumb += 1

		else:
			FirstSheet['A'+str(_LastColumb)] = "No Adjustments"
			FirstSheet['A'+str(_LastColumb)].style = styles.Style(font=Font(bold=False), alignment=Alignment(horizontal="center"))
			_LastColumb += 2

		FirstSheet['A' + str(_LastColumb)].style = styles.Style(border=Border(top=Side(style='thick')))
		FirstSheet['B' + str(_LastColumb)].style = styles.Style(border=Border(top=Side(style='thick')))
		FirstSheet['C' + str(_LastColumb)].style = styles.Style(border=Border(top=Side(style='thick')))

		_LastColumb += 1

		# -------------------------------------------------------------------------------
		# --------------------------------EMPLOYEES--------------------------------------
		# -------------------------------------------------------------------------------

		# get keys from working employee dictionary
		empKeys = unPickledData[1].keys()
		_Postitions = ["Line_Leader", "Line_Worker", "Mechanic"]

		#heres the methodolagy for itterating over employee dictionary
		for pos in _Postitions:

			FirstSheet['A'+str(_LastColumb)] = '----'+pos+'(s)----'
			FirstSheet['A'+str(_LastColumb)].style = styles.Style(font=Font(bold=True), alignment=Alignment(horizontal="left"))
			_LastColumb += 1

			for key in empKeys:
				if(unPickledData[1][key][0] == pos):

					try:
						int(key)
					except ValueError:
						FirstSheet['A' + str(_LastColumb)] = key
					else:
						FirstSheet['A' + str(_LastColumb)] = "Emloyee# " + key
					finally:
						FirstSheet['A' + str(_LastColumb)].style = styles.Style(font=Font(bold=True), alignment=Alignment(horizontal="center"))

					counterer = 1
					for badgeInTime, badgeOutTime in unPickledData[1][key][1]:
						if badgeOutTime is None:
							badgeOutTime = unPickledData[0]["Time Log Created"]

						#Only write the times we care about
						if badgeOutTime > unPickledData[0]["WO StartTime"]:

							if counterer > 1:
								FirstSheet['A' + str(_LastColumb)] = "-"
								FirstSheet['A' + str(_LastColumb)].style = styles.Style(font=Font(bold=False), alignment=Alignment(horizontal="center"))

							# Badge In Time
							FirstSheet["B" + str(_LastColumb)] = str(counterer) + " Badge in @: " + badgeInTime.strftime('%H:%M:%S')

							# Badge Out Time
							FirstSheet["C" + str(_LastColumb)] = str(counterer) + " Badge Out @: " + badgeOutTime.strftime('%H:%M:%S')

							counterer += 1
							_LastColumb += 1
			_LastColumb += 1

		FirstSheet['A' + str(_LastColumb)].style = styles.Style(border=Border(top=Side(style='thick')))
		FirstSheet['B' + str(_LastColumb)].style = styles.Style(border=Border(top=Side(style='thick')))
		FirstSheet['C' + str(_LastColumb)].style = styles.Style(border=Border(top=Side(style='thick')))

		_LastColumb += 1

		# -------------------------------------------------------------------------------
		# --------------------------------DOWNTIMES--------------------------------------
		# -------------------------------------------------------------------------------

		FirstSheet['A' + str(_LastColumb)] = '----Down Time----'
		FirstSheet['A' + str(_LastColumb)].style = styles.Style(font=Font(bold=True))
		_LastColumb += 1

		for DwnTime_Header, DwmTime_1Key, DwnTime_2Key in [ ("Maintanance > ", "FormattedMain", "Maintanance Down Times"),("Inventory > ", "FormattedInv", "Inventory Down Time"),("Quality_Control > ", "FormattedQuality", "Quality Control Down Time"),("Break > ","FormattedBreak", "Break Down Time"), ("ChangeOver > ","FormattedChngOvr","ChangeOver Time")]:

			FirstSheet['A' + str(_LastColumb)] = "" + DwnTime_Header + str(unPickledData[2][DwmTime_1Key][0])+':'+str(unPickledData[2][DwmTime_1Key][1])+':'+str(unPickledData[2][DwmTime_1Key][2])
			FirstSheet['A' + str(_LastColumb)].style = styles.Style(font=Font(bold=True), alignment=Alignment(horizontal="center"))
			_LastColumb += 1

			for (counter, (start,end)) in enumerate(unPickledData[0][DwnTime_2Key]):

				if end is None:
					end = (now, "000")

				FirstSheet["A"+str(_LastColumb)] = DwnTime_Header[0] + str(counter+1)+"("+str(end[0]-start[0]).split(".")[0]+"):"
				FirstSheet["A"+str(_LastColumb)].style = styles.Style(font=Font(bold=False), alignment=Alignment(horizontal="center"))

				#Place Start
				FirstSheet["B"+str(_LastColumb)] = 'Started by: '+str(start[1])+' @ '+start[0].strftime('%H:%M:%S')

				#Place End
				if not end == None:
					FirstSheet["C"+str(_LastColumb)] = ' Ended by: '+str(end[1])+' @ '+end[0].strftime('%H:%M:%S')
				else:
					FirstSheet["C"+str(_LastColumb)] = 'Ended by: <N/A> @ Machine Is Still Down'

			_LastColumb+=2

		FirstSheet['A'+str(_LastColumb)] = 'Total> '+str(unPickledData[2]["FormattedTotal"][0])+':'+str(unPickledData[2]["FormattedTotal"][1])+':'+str(unPickledData[2]["FormattedTotal"][2])
		FirstSheet['A'+str(_LastColumb)].style = styles.Style(font=Font(bold=True), alignment=Alignment(horizontal="center"))
		_LastColumb += 1

		FirstSheet['A' + str(_LastColumb)].style = styles.Style(border=Border(top=Side(style='thick')))
		FirstSheet['B' + str(_LastColumb)].style = styles.Style(border=Border(top=Side(style='thick')))
		FirstSheet['C' + str(_LastColumb)].style = styles.Style(border=Border(top=Side(style='thick')))

		_LastColumb += 2

		# -------------------------------------------------------------------------------
		# -----------------------------Batch&Pallets&QC----------------------------------
		# -------------------------------------------------------------------------------

		FirstSheet["A"+str(_LastColumb)] = "---Run Batch Info---"
		FirstSheet["A"+str(_LastColumb)].style= styles.Style(font=Font( bold=True, ))

		_RowLetter = "B"
		for batchHeaders in unPickledData[4]["INIT"]:
			FirstSheet[_RowLetter+str(_LastColumb)] = batchHeaders
			FirstSheet[_RowLetter+str(_LastColumb)].style = styles.Style(font=Font( bold=True, ), border=Border(bottom=Side(style='thick')), alignment=Alignment(horizontal="center"))
			_RowLetter = chr(ord(_RowLetter)+1)
		FirstSheet.row_dimensions[_LastColumb].height = 40
		_LastColumb += 1

		for batchInfoItems in sorted(unPickledData[4], key=unPickledData[4].get)[:-1]:
			_RowLetter = "B"
			FirstSheet["A"+str(_LastColumb)] = str(batchInfoItems)+":"
			FirstSheet["A"+str(_LastColumb)].style = styles.Style(font=Font(bold=True, ), border=Border(right=Side(style='thick')), alignment=Alignment(horizontal="right"))
			for batch2ndHeaders in unPickledData[4][batchInfoItems]:
				FirstSheet[_RowLetter+str(_LastColumb)] = batch2ndHeaders
				FirstSheet[_RowLetter+str(_LastColumb)].style = styles.Style(alignment=Alignment(horizontal="center"))
				_RowLetter = chr(ord(_RowLetter)+1)
			_LastColumb += 1
		_LastColumb += 1

		_LastColumb += 1
		FirstSheet["A"+str(_LastColumb)] = "---Run Pallet Info---"
		FirstSheet["A"+str(_LastColumb)].style= styles.Style(font=Font( bold=True, ))

		_RowLetter = "B"
		for palletHeaders in unPickledData[5]["INIT"]:
			FirstSheet[_RowLetter+str(_LastColumb)] = palletHeaders
			FirstSheet[_RowLetter+str(_LastColumb)].style = styles.Style(font=Font(bold=True, ), border=Border(bottom=Side(style='thick')),alignment=Alignment(horizontal="center"))
			_RowLetter = chr(ord(_RowLetter)+1)
		FirstSheet.row_dimensions[_LastColumb].height = 40
		_LastColumb+=1

		for palletInfoItems in sorted(unPickledData[5], key=unPickledData[4].get)[:-1]:
			_RowLetter = "B"
			FirstSheet["A"+str(_LastColumb)] = str(palletInfoItems)+":"
			FirstSheet["A"+str(_LastColumb)].style = styles.Style(font=Font(bold=True, ), border=Border(right=Side(style='thick')), alignment=Alignment(horizontal="right"))
			for batch2ndHeaders in unPickledData[5][palletInfoItems]:
				FirstSheet[_RowLetter+str(_LastColumb)] = batch2ndHeaders
				FirstSheet[_RowLetter+str(_LastColumb)].style = styles.Style(alignment=Alignment(horizontal="center"))
				_RowLetter = chr(ord(_RowLetter)+1)
			_LastColumb+=1
		_LastColumb+=1

		_LastColumb += 1
		FirstSheet["A"+str(_LastColumb)] = "---Run QC Info---"
		FirstSheet["A"+str(_LastColumb)].style= styles.Style(font=Font( bold=True, ))

		_RowLetter = "B"
		for QCHeaders in unPickledData[6]["INIT"]:
			FirstSheet[_RowLetter+str(_LastColumb)] = QCHeaders
			FirstSheet[_RowLetter+str(_LastColumb)].style = styles.Style(font=Font( bold=True, ), border=Border(bottom=Side(style='thick')), alignment=Alignment(horizontal="center"))
			_RowLetter = chr(ord(_RowLetter)+1)
		FirstSheet.row_dimensions[_LastColumb].height = 40
		_LastColumb += 1


		for QCInfoItems in sorted(unPickledData[6], key=unPickledData[4].get)[:-1]:
			_RowLetter = "B"
			FirstSheet["A"+str(_LastColumb)] = str(QCInfoItems)+":"
			FirstSheet["A"+str(_LastColumb)].style = styles.Style(font=Font(bold=True, ), border=Border(right=Side(style='thick')), alignment=Alignment(horizontal="right"))
			for batch2ndHeaders in unPickledData[6][QCInfoItems]:
				FirstSheet[_RowLetter+str(_LastColumb)] = batch2ndHeaders
				FirstSheet[_RowLetter+str(_LastColumb)].style = styles.Style(alignment=Alignment(horizontal="center"))
				_RowLetter = chr(ord(_RowLetter)+1)
			_LastColumb+=1
		_LastColumb+=1

		FirstSheet.column_dimensions["A"].width = 30.0
		FirstSheet.column_dimensions["B"].width = 30.0
		FirstSheet.column_dimensions["C"].width = 30.0
		FirstSheet.column_dimensions["D"].width = 30.0
		FirstSheet.column_dimensions["E"].width = 30.0
		FirstSheet.column_dimensions["F"].width = 30.0
		FirstSheet.column_dimensions["G"].width = 30.0
		# -------------------------------------------------------------------------------
		# --------------------------------SAVE&MSG---------------------------------------
		# -------------------------------------------------------------------------------

		wb.save(WO_LogFolder+str(w0)+".xlsx")#<<<<<<<<<<<<<<<<<<<<-------------------------------------------- Save the file
		print "Saved new Log: ", WO_LogFolder+str(w0)+".xlsx", " Run: ", curNumRuns+1


	def writeToSQL(self, MachineLog, databaseConectionVars=("192.168.20.31", "cyrus", "cyrus2sql", "pqmfg_daq")):
		sql = []
		NOW = datetime.datetime.now()

		# Open database connection
		db = MySQLdb.connect(databaseConectionVars[0], databaseConectionVars[1], databaseConectionVars[2], databaseConectionVars[3])

		db.query("""select * from WORKORDER_RUNS where WORKORDER_NUM = """ + MachineLog[0]["WO"])
		RunNum = str(int(db.store_result().num_rows()) + 1)

		# -------------------------------------------------------------------------------
		# --------------------------WORKORDER_RUNS---------------------------------------
		# -------------------------------------------------------------------------------

		# woRunSQLVars = ["WORKORDER_NUM", "RUN_NUM","MACHINE_NUM", "RUN_START", "RUN_END", "FILL_START", "FILL_END",
		# 							"TOTAL_COUNT", "TOTAL_BOXED", "TOTAL_SCRAPPED", "TARE_WEIGHT", "VOLUME", "SPECIFIC_GRAVITY",
		# 							"WEIGHT", "Cosmetic", "ITEM_NUMBER", "DESIRED_QTY", "FORMULA_REF_NUM","PACKING_CODE", "PACK_OFF",
		# 							"PUMP_NUM", "SIMPLEX_NUM"]

		# ---------------------------------ZIP---------------------------------------

		machineQuery = ["WORKORDER_NUM", "MACHINE_NUM", "RUN_NUM", "RUN_START",
										"RUN_END", "FILL_START", "FILL_END", "TOTAL_COUNT",
										"TOTAL_BOXED", "TOTAL_SCRAPPED"]

		machineDictKeys = ["WO", "Machine ID", "run_num", "WO StartTime", "Time Log Created",
											"Fill Start", "Fill End", "Total Count", "Box Count",
											"Fail Count"]

		# ---------------------------------ZIP2--------------------------------------

		fillSheetQuery = ["TARE_WEIGHT", "VOLUME", "SPECIFIC_GRAVITY", "WEIGHT",
							"Cosmetic", "ITEM_NUMBER", "DESIRED_QTY", "FORMULA_REF_NUM",
							"PACKING_CODE", "PACK_OFF", "PUMP_NUM", "SIMPLEX_NUM"]

		fillSheetDictKeys = ["Tare Weight(g)", "Volume(ml)", "    Specific Gravity",
												"Weight(g)", "Cosmetic", "Item Number", "Desired Qty",
												"Formula Ref#", "Packing Code", "Pack Off", "Pump#",
												"Simplex#"]

		VarHeaders = ""
		EndingString = ""

		for MV_DicKey, MV_SQLKey in zip(machineDictKeys, machineQuery):

			if MV_DicKey in MachineLog[0].keys() or MV_SQLKey == "RUN_NUM":
				VarHeaders += MV_SQLKey + ", "
				if not MV_SQLKey == "RUN_NUM" and MachineLog[0][MV_DicKey] == None:
					EndingString += "'NULL', "
				else:
					if MV_SQLKey.endswith("_END") or MV_SQLKey.endswith("_START"):

						if MachineLog[0][MV_DicKey] is not None:
							EndingString += "'" + MachineLog[0][MV_DicKey].strftime('%Y-%m-%d %H:%M:%S') + "', "

						else:
							EndingString += "'" + str(MachineLog[0][MV_DicKey]) + "', "

					elif MV_SQLKey == "RUN_NUM":
						EndingString += "'" + str(RunNum) + "', "

					elif MV_SQLKey == "TOTAL_BOXED" or MV_SQLKey == "TOTAL_COUNT":
						EndingString += "'" + str(sum(MachineLog[0][MV_DicKey])) + "', "

					else:
						EndingString += "'" + str(MachineLog[0][MV_DicKey]) + "', "

		for FS_DicKey, FS_SQLKey in zip(fillSheetDictKeys, fillSheetQuery):
			if FS_DicKey in MachineLog[3].keys():
				VarHeaders += FS_SQLKey + ", "
				if MachineLog[3][FS_DicKey] == None:
					EndingString += "'NULL', "
				else:
					EndingString += "'" + str(MachineLog[3][FS_DicKey]) + "', "

		sql.append("INSERT INTO WORKORDER_RUNS (" + VarHeaders[:-2] + ") VALUES (" + EndingString[:-2] + ");")

		# -------------------------------------------------------------------------------
		# --------------------------DOWNTIMES--------------------------------------------
		# -------------------------------------------------------------------------------

		# 1) Maitenance, 2) Inventory, 3) Quality_Control 4) Break 5) ChangeOver
		# 'ChangeOver','Maitenance','Inventory','Quality_Control','Break'
		# WORKORDER_NUM, RUN_NUM, TYPE, START, END, EMP_BD, EMP_BU
		VarHeaders = "WORKORDER_NUM, MACHINE_NUM, RUN_NUM, TYPE, START, END, EMP_BD, EMP_BU"
		EndingString = ""

		ChangeDwnTime = ("ChangeOver", MachineLog[0]["ChangeOver Time"])
		MainDwnTime = ("Maitenance", MachineLog[0]["Maintanance Down Times"])
		InvDwnTime = ("Inventory", MachineLog[0]["Inventory Down Time"])
		QualDwnTime = ("Quality_Control", MachineLog[0]["Quality Control Down Time"])
		BreakDwnTime = ("Break", MachineLog[0]["Break Down Time"])

		for reason, DWN_Tmes in [ChangeDwnTime, MainDwnTime, InvDwnTime, QualDwnTime, BreakDwnTime]:

			for DWN_Tme in DWN_Tmes:

				if DWN_Tme[1] is not None:
					sql.append("INSERT INTO DOWNTIMES (" + VarHeaders + ") VALUES ( '%s','%s','%s','%s','%s','%s','%s','%s');" % (MachineLog[0]["WO"], MachineLog[0]["Machine ID"], str(RunNum), reason, DWN_Tme[0][0].strftime('%Y-%m-%d %H:%M:%S'), DWN_Tme[1][0].strftime('%Y-%m-%d %H:%M:%S'), DWN_Tme[0][1], DWN_Tme[1][1]))
				else:
					sql.append("INSERT INTO DOWNTIMES (" + VarHeaders + ") VALUES ( '%s','%s','%s','%s','%s','%s','%s','%s');" % (MachineLog[0]["WO"], MachineLog[0]["Machine ID"], str(RunNum), reason, DWN_Tme[0][0].strftime('%Y-%m-%d %H:%M:%S'), NOW.strftime('%Y-%m-%d %H:%M:%S'), DWN_Tme[0][1], "000"))

		# -------------------------------------------------------------------------------
		# --------------------------EMPLOYEE_BADGE_SWIPES--------------------------------
		# -------------------------------------------------------------------------------

		VarHeaders = "EMPLOYEE_BADGE_NUM, EMP_TYPE, MACHINE_NUM, WORKORDER_NUM, RUN_NUM, TIME_IN, TIME_OUT"

		for key in MachineLog[1].keys():
			for times in MachineLog[1][key][1]:

				starttime = times[0].strftime('%Y-%m-%d %H:%M:%S')
				endtime = times[1]

				if endtime is not None:
					endtime = endtime.strftime('%Y-%m-%d %H:%M:%S')
				else:
					endtime = NOW.strftime('%Y-%m-%d %H:%M:%S')

				sql.append("INSERT INTO EMPLOYEE_BADGE_SWIPES (" + VarHeaders + ") VALUES ('%s','%s','%s','%s','%s','%s','%s');" % (key, MachineLog[1][key][0], MachineLog[0]["Machine ID"], MachineLog[0]["WO"], str(RunNum), starttime, endtime))

		# -------------------------------------------------------------------------------
		# -----------------------------------PALLETS-------------------------------------
		# -------------------------------------------------------------------------------
		if len(MachineLog[5].keys()) > 1:
			VarHeaders = "PALLET_NUM, BATCH_NUM, WORKORDER_NUM, RUN_NUM, BOXES, PEACES_PER_BOX"

			for key in MachineLog[5].keys():
				if not type(key) is str:
					sql.append("INSERT INTO PALLETS (" + VarHeaders + ") VALUES ('%s','%s','%s','%s','%s','%s');" % (MachineLog[5][key][0], MachineLog[5][key][4], MachineLog[0]["WO"], str(RunNum), MachineLog[5][key][1], MachineLog[5][key][2]))

		# -------------------------------------------------------------------------------
		# -----------------------------------BATCHES-------------------------------------
		# -------------------------------------------------------------------------------
		if len(MachineLog[4].keys()) > 1:
			VarHeaders = "BATCH_NUM, WORKORDER_NUM, MACHINE_NUM, RUN_NUM, FILL_WEIGHT, TOTAL_WEIGHT, TOTAL_WEIGHT_RANGE"

			for key in MachineLog[4].keys():
				if not type(key) is str:
					sql.append("INSERT INTO BATCHES (" + VarHeaders + ") VALUES ('%s','%s','%s','%s','%s','%s','%s');" % (MachineLog[4][key][0], MachineLog[0]["WO"], MachineLog[0]["Machine ID"], str(RunNum), MachineLog[4][key][1], MachineLog[4][key][2], MachineLog[4][key][3]))

		# -------------------------------------------------------------------------------
		# ------------------------------------QC-----------------------------------------
		# -------------------------------------------------------------------------------
		if len(MachineLog[6].keys()) > 1:
			VarHeaders = "MACHINE_NUM, WORKORDER_NUM, RUN_NUM, BATCH_NUM, STABILITY, BEGINS, MIDDLE, ENDS, RESAMPLE, INITIALS"

			for key in MachineLog[6].keys():
				if not type(key) is str:
					sql.append("INSERT INTO QC (" + VarHeaders + ") VALUES ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s');" % (MachineLog[0]["Machine ID"], MachineLog[0]["WO"], str(RunNum), MachineLog[6][key][0], MachineLog[6][key][1], MachineLog[6][key][2], MachineLog[6][key][3], MachineLog[6][key][4], MachineLog[6][key][5], MachineLog[6][key][6]))

		# -------------------------------------------------------------------------------
		# ----------------------------------FINALLY--------------------------------------
		# -------------------------------------------------------------------------------

		# prepare a cursor object using cursor() method
		cursor = db.cursor()

		for SQL_Statment in sql:
			try:
				# Execute the SQL command
				print "SQL: ", SQL_Statment
				cursor.execute(SQL_Statment)
				# Commit your changes in the database
				db.commit()

			except _mysql.Error, e:
				db.rollback()
				print "SQL Error %d: %s" % (e.args[0], e.args[1])

			except:
				# Rollback in case there is any error
				db.rollback()

		# disconnect from server
		db.close()

	def writeFile(self, FileName, Content , write ="w"):
		myfile = open(FileName, write)
		for newLine in Content:
			myfile.write(newLine+'\n')
		myfile.close()

	def stop(self):
		#set runflag to False

		if self.testing: print "Setting RunFlag to False"
		self.running = False

		#if self.testing: print "Waiting for alive thread to join"
		#self.IsMachineAlive.join()

		#Create a connection to self so that we skip of blocking call
		if self.testing: print "Creating Kill Connection"
		socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect(self.Addr)

		#Kill the pipe
		if self.testing: print "Clossing Pipe"
		self.serversock.close()


	def run(self):

		#All this loop does is listen for connections and spawn mini threads
		while self.running:

			#Here we wait for incoming connection
			clientsock, addr = self.serversock.accept()
			if self.testing: print "Connection Started: ", addr
			#we spawn new mini thread and pass off connection
			Thread(target=self.miniThread, args=(clientsock, addr)).start()

if __name__=='__main__':
	a = ThreadedTCPNetworkAgent(5006, Folder="/media/windowsshare/Operations/WorkOrderExcelLogs")
	a.start()